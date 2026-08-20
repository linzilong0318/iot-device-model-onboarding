#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 NJBK8 输入 Excel 生成 points.json(四维度点位清单)
- POWER_PARAMETER 测量数据 -> MeasurePoint
- DEVICE_SET_PARAMETER 保护数据 -> Attribute(保护阈值/使能/延时配置,readwrite)
- CONTROL_PARAMETER 控制命令 -> Service(writeonly 控制下发)
- event_log 保护事件 -> Event(故障/告警/操作/状态变化记录)
- SYSTEM_PARAMETER 运行数据 -> MeasurePoint(运行时间/次数/状态) + Attribute(额定参数/通讯/DI/DO 配置)
- MANUFACTURER_PARAMETER 厂内参数 -> Attribute(厂内配置,标注 inferred,一般不入模型)
"""
import json, re
from pathlib import Path
import openpyxl

BASE = Path(r"d:\项目\iot-device-model-onboarding")
SRC = BASE / "input" / "【马达保护器】NJBK8数据定义.xlsx"
OUT = BASE / "output" / "NJBK8" / "points.json"

# C 类型 -> 平台数据类型映射
DTYPE_MAP = {
    "uint8_t": "INT", "uint16_t": "INT", "uint32_t": "INT",
    "int8_t": "INT", "int16_t": "INT", "int32_t": "INT",
    "float": "FLOAT", "double": "FLOAT",
    "bool": "BOOL", "char": "STRING",
}

# 寄存器数据类型(C -> 点表 dataType)
REG_DTYPE_MAP = {
    "uint8_t": "u8", "uint16_t": "u16", "uint32_t": "u32",
    "int8_t": "i8", "int16_t": "i16", "int32_t": "i32",
    "float": "float32", "double": "double64",
}


def norm_coeff(c):
    if c is None or c == "" or c == " -":
        return None
    try:
        f = float(c)
        return f if f != 0 else None
    except Exception:
        return None


def norm_unit(u):
    if u is None or u == "" or u == " -":
        return ""
    return str(u).strip()


def norm_rw(rw):
    if not rw:
        return "R"
    s = str(rw).lower()
    if "write" in s and "read" in s:
        return "RW"
    if "write" in s:
        return "W"
    return "R"


def build_modbus_index(wb):
    """Modbus对象字典 -> {var_name: {addr, dtype, dlen, rw}}"""
    ws = wb["Modbus对象字典"]
    idx = {}
    cur_struct = None
    for r in range(2, ws.max_row + 1):
        struct = ws.cell(row=r, column=1).value
        var = ws.cell(row=r, column=3).value
        addr = ws.cell(row=r, column=5).value
        dtype = ws.cell(row=r, column=6).value
        dlen = ws.cell(row=r, column=7).value
        rw = ws.cell(row=r, column=8).value
        if struct:
            cur_struct = struct
        if var is None and struct is None:
            continue
        key = f"{cur_struct}::{var}"
        idx[key] = {
            "struct": cur_struct, "addr": addr, "dtype": dtype,
            "dlen": dlen, "rw": rw,
        }
    return idx


def parse_remark_enum(remark):
    """从备注解析枚举值,如 '0：禁用  1：跳闸 2：报警' -> {0:'禁用',1:'跳闸',2:'报警'}"""
    if not remark or remark == " -":
        return None
    s = str(remark)
    # 匹配 数字：文本 或 数字:文本
    pairs = re.findall(r'(\d+)\s*[：:]\s*([^\d：:]+?)(?=\s*\d+\s*[：:]|$)', s)
    if len(pairs) >= 2:
        out = {}
        for k, v in pairs:
            v = v.strip().rstrip(' 　')
            if v:
                out[int(k)] = v
        return out if out else None
    return None


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    mb_idx = build_modbus_index(wb)

    # 读取数据定义 sheet
    ws = wb["数据定义"]
    records = []
    cur_struct, cur_desc = None, None
    for r in range(2, ws.max_row + 1):
        struct = ws.cell(row=r, column=1).value
        struct_desc = ws.cell(row=r, column=2).value
        var = ws.cell(row=r, column=3).value
        var_desc = ws.cell(row=r, column=4).value
        dtype = ws.cell(row=r, column=5).value
        coeff = ws.cell(row=r, column=6).value
        dlen = ws.cell(row=r, column=7).value
        key = ws.cell(row=r, column=8).value
        unit = ws.cell(row=r, column=12).value
        remark = ws.cell(row=r, column=13).value
        if struct:
            cur_struct = struct
            cur_desc = struct_desc
        if var is None and struct is None:
            continue
        # 跳过分隔行(变量名为 ………)
        if var and "…" in str(var):
            continue
        records.append({
            "struct": cur_struct, "struct_desc": cur_desc,
            "var": var, "var_desc": (str(var_desc).strip() if var_desc else ""),
            "dtype": dtype, "coeff": norm_coeff(coeff),
            "dlen": dlen, "key": key, "unit": norm_unit(unit),
            "remark": remark if remark and remark != " -" else None,
        })

    # 四维度分组
    points = {"Attribute": [], "MeasurePoint": [], "Event": [], "Service": []}

    # ---------- POWER_PARAMETER -> MeasurePoint ----------
    for rec in [x for x in records if x["struct"] == "POWER_PARAMETER"]:
        mb = mb_idx.get(f"POWER_PARAMETER::{rec['var']}", {})
        enum = parse_remark_enum(rec["remark"])
        points["MeasurePoint"].append({
            "source_var": rec["var"], "name": rec["var_desc"],
            "platform_dtype": DTYPE_MAP.get(rec["dtype"], "FLOAT"),
            "unit": rec["unit"], "rw": "R",
            "enum": enum, "remark": rec["remark"],
            "reg_dtype": REG_DTYPE_MAP.get(rec["dtype"], "u16"),
            "reg_addr": mb.get("addr"), "reg_count": mb.get("dlen"),
            "reg_rw": norm_rw(mb.get("rw")), "coeff": rec["coeff"],
            "struct": "POWER_PARAMETER", "inferred": False,
        })

    # ---------- SYSTEM_PARAMETER:运行统计 -> MeasurePoint;配置 -> Attribute ----------
    # 运行统计类测点(只读,运行时间/次数/最大值)
    sys_measure_keys = {
        "run_time_now", "stop_time_now", "run_time_all", "stop_time_all",
        "start_number", "trip_number", "start_current_max", "run_current_max",
        "work_mode",
    }
    for rec in [x for x in records if x["struct"] == "SYSTEM_PARAMETER"]:
        mb = mb_idx.get(f"SYSTEM_PARAMETER::{rec['var']}", {})
        enum = parse_remark_enum(rec["remark"])
        if rec["var"] in sys_measure_keys:
            points["MeasurePoint"].append({
                "source_var": rec["var"], "name": rec["var_desc"],
                "platform_dtype": DTYPE_MAP.get(rec["dtype"], "FLOAT"),
                "unit": rec["unit"], "rw": "R",
                "enum": enum, "remark": rec["remark"],
                "reg_dtype": REG_DTYPE_MAP.get(rec["dtype"], "u16"),
                "reg_addr": mb.get("addr"), "reg_count": mb.get("dlen"),
                "reg_rw": norm_rw(mb.get("rw")), "coeff": rec["coeff"],
                "struct": "SYSTEM_PARAMETER", "inferred": False,
            })
        else:
            # 配置类属性(额定参数/通讯/DI/DO 配置)
            points["Attribute"].append({
                "source_var": rec["var"], "name": rec["var_desc"],
                "platform_dtype": "ENUM" if enum else DTYPE_MAP.get(rec["dtype"], "STRING"),
                "unit": rec["unit"], "is_required": "False",
                "enum": enum, "remark": rec["remark"],
                "reg_dtype": REG_DTYPE_MAP.get(rec["dtype"], "u16"),
                "reg_addr": mb.get("addr"), "reg_count": mb.get("dlen"),
                "reg_rw": norm_rw(mb.get("rw")), "coeff": rec["coeff"],
                "struct": "SYSTEM_PARAMETER", "inferred": False,
            })

    # ---------- DEVICE_SET_PARAMETER -> Attribute(保护配置) ----------
    for rec in [x for x in records if x["struct"] == "DEVICE_SET_PARAMETER"]:
        mb = mb_idx.get(f"DEVICE_SET_PARAMETER::{rec['var']}", {})
        enum = parse_remark_enum(rec["remark"])
        points["Attribute"].append({
            "source_var": rec["var"], "name": rec["var_desc"],
            "platform_dtype": "ENUM" if enum else DTYPE_MAP.get(rec["dtype"], "INT"),
            "unit": rec["unit"], "is_required": "False",
            "enum": enum, "remark": rec["remark"],
            "reg_dtype": REG_DTYPE_MAP.get(rec["dtype"], "u16"),
            "reg_addr": mb.get("addr"), "reg_count": mb.get("dlen"),
            "reg_rw": norm_rw(mb.get("rw")), "coeff": rec["coeff"],
            "struct": "DEVICE_SET_PARAMETER", "inferred": False,
        })

    # ---------- MANUFACTURER_PARAMETER -> Attribute(厂内参数,标注 inferred) ----------
    for rec in [x for x in records if x["struct"] == "MANUFACTURER_PARAMETER"]:
        # 跳过校表命令类(变量名含函数签名)
        if rec["var"] and ("(" in str(rec["var"]) or str(rec["var"]).startswith("device.")):
            continue
        mb = mb_idx.get(f"MANUFACTURER_PARAMETER::{rec['var']}", {})
        enum = parse_remark_enum(rec["remark"])
        points["Attribute"].append({
            "source_var": rec["var"], "name": rec["var_desc"],
            "platform_dtype": "ENUM" if enum else DTYPE_MAP.get(rec["dtype"], "STRING"),
            "unit": rec["unit"], "is_required": "False",
            "enum": enum, "remark": rec["remark"],
            "reg_dtype": REG_DTYPE_MAP.get(rec["dtype"], "u16"),
            "reg_addr": mb.get("addr"), "reg_count": mb.get("dlen"),
            "reg_rw": norm_rw(mb.get("rw")), "coeff": rec["coeff"],
            "struct": "MANUFACTURER_PARAMETER", "inferred": True,
        })

    # ---------- CONTROL_PARAMETER -> Service ----------
    service_map = {
        "start1": ("Start1", "起动1"),
        "start2": ("Start2", "起动2"),
        "stop": ("Stop", "停机"),
        "reset": ("Reset", "复位"),
        "powerclear": ("PowerClear", "电量清零"),
        "eventclear": ("EventClear", "事件总清"),
        "operationclear": ("OperationClear", "运行信息清零"),
        "reset_run_timeout": ("ResetRunTimeout", "复位运行超时"),
        "reset_fault_number": ("ResetFaultNumber", "复位故障次数"),
        "flashclear": ("FlashClear", "恢复出厂"),
    }
    for rec in [x for x in records if x["struct"] == "CONTROL_PARAMETER"]:
        mb = mb_idx.get(f"CONTROL_PARAMETER::{rec['var']}", {})
        en_id, en_name = service_map.get(rec["var"], (rec["var"], rec["var_desc"]))
        points["Service"].append({
            "source_var": rec["var"], "name": rec["var_desc"],
            "suggested_id": en_id, "suggested_name": en_name,
            "reg_addr": mb.get("addr"), "reg_rw": norm_rw(mb.get("rw")),
            "struct": "CONTROL_PARAMETER", "inferred": False,
        })

    # ---------- event_log -> Event ----------
    # event_log 是历史记录(上1次/上10次),按记录类型分组归纳为事件
    # 故障记录、告警记录、操作记录、状态变化记录
    # 故障原因枚举(从 fault_data remark 提取)
    fault_remark = None
    for rec in [x for x in records if x["struct"] == "event_log" and x["var"] == "fault_data"]:
        if "故障原因" in rec["var_desc"] or "跳闸" in rec["var_desc"]:
            fault_remark = rec["remark"]
            break
    fault_enum = parse_remark_enum(fault_remark) if fault_remark else None

    # 归纳 4 类事件
    events_def = [
        ("TripRecord", "跳闸记录", "FAULT", "故障跳闸记录(含故障原因/相别/时刻及跳闸前电气量)"),
        ("AlarmRecord", "告警记录", "ALARM", "告警记录(含告警原因/相别/时刻及告警前电气量)"),
        ("OperationRecord", "操作记录", "INFO", "操作记录(含操作方式/发生时刻)"),
        ("StateChangeRecord", "状态变化记录", "INFO", "DI/DO 状态变化记录(SOE)"),
    ]
    for eid, ename, etype, desc in events_def:
        points["Event"].append({
            "suggested_id": eid, "name": ename,
            "event_type": etype, "desc": desc,
            "fault_enum": fault_enum if etype in ("FAULT", "ALARM") else None,
            "struct": "event_log", "inferred": True,
        })

    # 统计
    summary = {k: len(v) for k, v in points.items()}
    out = {
        "device": {"name": "NJBK8", "type": "马达保护器",
                   "protocol": "Modbus",
                   "source_file": str(SRC.name)},
        "summary": summary,
        "points": points,
    }
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print("points.json 已生成:", OUT)
    print("四维度点位数:", summary)
    print(f"  MeasurePoint: {summary['MeasurePoint']} (POWER_PARAMETER 66 + SYSTEM_PARAMETER 运行 9)")
    print(f"  Attribute: {summary['Attribute']} (DEVICE_SET_PARAMETER 保护配置 + SYSTEM_PARAMETER 配置 + MANUFACTURER 厂内)")
    print(f"  Event: {summary['Event']} (event_log 归纳 4 类)")
    print(f"  Service: {summary['Service']} (CONTROL_PARAMETER 10)")


if __name__ == "__main__":
    main()
