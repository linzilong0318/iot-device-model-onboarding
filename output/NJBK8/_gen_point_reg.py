#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成 point_reg.json(点表寄存器映射)
模型 pointKey -> Modbus 寄存器信息(address/registerCount/functionCode/dataType/coefficient)
pointKey 全集 = 设备模型测点(FromDeviceType 引用 39 + 子表新增 17)
"""
import json
from pathlib import Path
import openpyxl

BASE = Path(r"d:\项目\iot-device-model-onboarding")
SRC = BASE / "input" / "【马达保护器】NJBK8数据定义.xlsx"
MODEL_SPEC = BASE / "output" / "NJBK8" / "model_spec.json"
POINTS = BASE / "output" / "NJBK8" / "points.json"
OUT = BASE / "output" / "NJBK8" / "point_reg.json"

# 模型 pointKey -> 输入 Excel 变量名 的映射
# 引用测点(类型标识符)-> 输入变量名
REF_KEY_TO_VAR = {
    "Uab": "Total_AB_Voltage", "Ubc": "Total_BC_Voltage", "Uca": "Total_AC_Voltage",
    "Ua": "LA_result->Voltage", "Ub": "LB_result->Voltage", "Uc": "LC_result->Voltage",
    "Un": "Voltage_N",
    "Ia": "LA_result->Current", "Ib": "LB_result->Current", "Ic": "LC_result->Current",
    "LeakCurrent": "Total_Creepage", "GroundCurrent": "Current_Ground",
    "EP": "Total_Electric_Quantity",
    "EPI_PhaseA": "LA_result->Active_Energy", "EPI_PhaseB": "LB_result->Active_Energy", "EPI_PhaseC": "LC_result->Active_Energy",
    "EQI_PhaseA": "LA_result->Reactive_Energy", "EQI_PhaseB": "LB_result->Reactive_Energy", "EQI_PhaseC": "LC_result->Reactive_Energy",
    "Pa": "LA_result->Active_Power[4]", "Pc": "LC_result->Active_Power[4]",
    "Qa": "LA_result->Reactive_Power[4]", "Qb": "LB_result->Reactive_Power[4]", "Qc": "LC_result->Reactive_Power[4]",
    "Sa": "LA_result->Apparent_Power[4]", "Sb": "LB_result->Apparent_Power[4]", "Sc": "LC_result->Apparent_Power[4]",
    "PFa": "LA_result->Power_Factor", "PFb": "LB_result->Power_Factor", "PFc": "LC_result->Power_Factor",
    "PhaseAngleA": "LA_result->Angle[0]", "PhaseAngleB": "LB_result->Angle[0]", "PhaseAngleC": "LC_result->Angle[0]",
    "FundamentalVa": "LA_result->Voltage_wave", "FundamentalVb": "LB_result->Voltage_wave", "FundamentalVc": "LC_result->Voltage_wave",
    "FundamentalIa": "LA_result->Current_wave", "FundamentalIb": "LB_result->Current_wave", "FundamentalIc": "LC_result->Current_wave",
}
# 新增测点(自定义标识符)-> 输入变量名
ADD_KEY_TO_VAR = {
    "RunTimeNow": "run_time_now", "RunTimeAll": "run_time_all",
    "StartNumber": "start_number", "TripNumber": "trip_number",
    "StartCurrentMax": "start_current_max", "RunCurrentMax": "run_current_max",
    "WorkMode": "work_mode",
    "DI0": "DI_state[0]", "DI1": "DI_state[1]", "DI2": "DI_state[2]",
    "DI3": "DI_state[3]", "DI4": "DI_state[4]", "DI5": "DI_state[5]",
    "DO0": "DO_state[0]", "DO1": "DO_state[1]", "DO2": "DO_state[2]", "DO3": "DO_state[3]",
}
KEY_TO_VAR = {**REF_KEY_TO_VAR, **ADD_KEY_TO_VAR}

# C 类型 -> 点表 dataType
REG_DTYPE_MAP = {
    "uint8_t": "u8", "uint16_t": "u16", "uint32_t": "u32",
    "int8_t": "i8", "int16_t": "i16", "int32_t": "i32",
    "float": "float32", "double": "double64",
}


def build_modbus_index(wb):
    """Modbus对象字典 -> {struct::var: {addr, dtype, dlen, rw}}"""
    ws = wb["Modbus对象字典"]
    idx = {}
    cur_struct = None
    for r in range(2, ws.max_row + 1):
        struct = ws.cell(row=r, column=1).value
        var = ws.cell(row=r, column=3).value
        addr = ws.cell(row=r, column=5).value
        dtype = ws.cell(row=r, column=6).value
        dlen = ws.cell(row=r, column=7).value
        rw = ws.cell(row=r, column=8).value
        if struct:
            cur_struct = struct
        if var is None and struct is None:
            continue
        if var and "…" in str(var):
            continue
        idx[f"{cur_struct}::{var}"] = {
            "addr": addr, "dtype": dtype, "dlen": dlen, "rw": rw,
        }
    return idx


def build_data_def_index(wb):
    """数据定义 sheet -> {struct::var: {coeff, unit}}"""
    ws = wb["数据定义"]
    idx = {}
    cur_struct = None
    for r in range(2, ws.max_row + 1):
        struct = ws.cell(row=r, column=1).value
        var = ws.cell(row=r, column=3).value
        coeff = ws.cell(row=r, column=6).value
        unit = ws.cell(row=r, column=12).value
        if struct:
            cur_struct = struct
        if var is None and struct is None:
            continue
        if var and "…" in str(var):
            continue
        c = None
        if coeff and coeff != " -":
            try:
                c = float(coeff)
                if c == 0:
                    c = None
            except Exception:
                c = None
        u = ""
        if unit and unit != " -":
            u = str(unit).strip()
        idx[f"{cur_struct}::{var}"] = {"coeff": c, "unit": u}
    return idx


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    mb_idx = build_modbus_index(wb)
    dd_idx = build_data_def_index(wb)

    spec = json.loads(MODEL_SPEC.read_text(encoding="utf-8"))
    # 模型测点全集 = select.MeasurePoint + add.MeasurePoint 的 *ID
    model_mp_keys = list(spec["select"]["MeasurePoint"])
    for r in spec["add"]["MeasurePoint"]:
        model_mp_keys.append(r["*ID"])

    rows = {}
    not_found = []
    for pk in model_mp_keys:
        var = KEY_TO_VAR.get(pk)
        if not var:
            not_found.append(pk)
            continue
        # 在 Modbus 索引中查找(需遍历所有 struct 找匹配 var)
        mb = None
        for k, v in mb_idx.items():
            if k.endswith(f"::{var}"):
                mb = v
                break
        if not mb or not mb.get("addr"):
            not_found.append(pk)
            continue
        dd = None
        # 数据定义 sheet 变量名无 [4] 后缀,去掉后缀匹配
        var_base = var.replace("[4]", "")
        for k, v in dd_idx.items():
            if k.endswith(f"::{var}") or k.endswith(f"::{var_base}"):
                dd = v
                break
        # functionCode: readonly -> [03,00], readwrite -> [03,06], writeonly -> [00,06]
        rw = (mb.get("rw") or "").lower()
        if "write" in rw and "read" in rw:
            fc = "[03,06]"
        elif "write" in rw:
            fc = "[00,06]"
        else:
            fc = "[03,00]"
        # registerCount: 数据长度(字节数 / 2 = 寄存器数);uint32/float=2, uint16=1
        dlen = mb.get("dlen")
        try:
            dlen_int = int(dlen) if dlen else 1
            reg_count = dlen_int // 2 if dlen_int >= 2 else 1
        except Exception:
            reg_count = 1
        rows[pk] = {
            "address": mb["addr"],
            "registerCount": reg_count,
            "functionCode": fc,
            "dataType": REG_DTYPE_MAP.get(mb.get("dtype"), "u16"),
            "coefficient": dd["coeff"] if dd else None,
            "unit": dd["unit"] if dd else "",
        }

    point_reg = {
        "protocol": "ModbusRTU_Vega_ARM64_V1.1.0",
        "template": str(BASE / "templates" / "ModbusRTU_Vega_ARM64_V1.1.0.xlsx"),
        "model_xlsx": str(BASE / "output" / "NJBK8" / "NJBK8_设备模型_20260820.xlsx"),
        "output": str(BASE / "output" / "NJBK8" / "NJBK8_点表_20260820.xlsx"),
        "rows": rows,
    }
    OUT.write_text(json.dumps(point_reg, ensure_ascii=False, indent=2), encoding="utf-8")
    print("point_reg.json 已生成:", OUT)
    print(f"模型测点全集: {len(model_mp_keys)}")
    print(f"点表行(有寄存器信息): {len(rows)}")
    print(f"未找到地址(不生成行): {len(not_found)}: {not_found}")


if __name__ == "__main__":
    main()
