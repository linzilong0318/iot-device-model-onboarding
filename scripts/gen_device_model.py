#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成设备模型 Excel(分支 A)—— 从设备类型筛选点位 + 私有新增点位

输入:
  model_spec.json  —— 设备模型配置(模型元信息 + SELECT 引用清单 + ADD 新增点位)
  raw 物模型文档   —— 引用点位的字段来源(类型标准定义)

输出:
  按模板 model_template.xlsx 结构生成的设备模型 xlsx

用法:
  python gen_device_model.py <model_spec.json> [输出路径覆盖]
  (model_spec.json 内可指定 raw_doc / template / output 路径,命令行参数优先)

铁律:
  - FromDeviceType 子表 = 从类型引用的点位清单(匹配上的类型点位);
    四张子表(Attribute/MeasurePoint/Event/Service)只能写 FromDeviceType 中
    未引用的【新增点位】,从类型引用的点位绝不重复写入四张子表。
  - 引用点位字段取自类型(raw 文档)标准定义;新增点位字段取自 spec(类型中无此点位)。
  - 生成物 ID 一律 project_ 前缀(public_ 仅用于引用知识库公有类型)。
  - 引用服务/事件的 Input/Output 依赖点位,脚本 ensure_refs_present 自动补入引用。

model_spec.json 结构:
{
  "raw_doc": "/opt/data/wiki/raw/papers/public_ElectricMeter_3P_V1_0_2.md",
  "template": "templates/model_template.xlsx",
  "output": "/opt/data/output/xxx_设备模型_yyyyMMdd.xlsx",
  "model": {
    "id": "project_xxx", "name": "...", "name_zh": "...", "name_en": "...",
    "device_type": "public_ElectricMeter_3P_V1_0_2"
  },
  "select": {
    "Attribute": ["SN", "ProductCategory"],
    "MeasurePoint": ["Ua", "Ub"],
    "Event": ["AlarmRevU"],
    "Service": ["ClearECmd"]
  },
  "add": {
    "Attribute": [{"*ID": "Net", "*Name_default": "...", "*DataType": "ENUM",
                   "DataDefine": "{...}", "Unit": "", "*IsRequired": "False", "Desc": "..."}],
    "MeasurePoint": [...],
    "Event": [],
    "Service": []
  }
}
"""
import io, json, os, re, sys
import openpyxl

DIMS = ['Attribute', 'MeasurePoint', 'Event', 'Service']
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def parse_raw(path):
    """解析 raw 物模型文档 -> {维度: [表头, 行1, 行2, ...]}"""
    txt = io.open(path, encoding='utf-8').read()
    sections, cur = {}, None
    for line in txt.splitlines():
        m = re.match(r'^## (\w+)$', line.strip())
        if m:
            cur = m.group(1)
            sections[cur] = []
        elif cur and line.startswith('|') and not line.startswith('|---'):
            cells = [c.strip() for c in line.strip().strip('|').split('|')]
            if cells and cells[0] and cells[0] != '---':
                sections[cur].append(cells)
    return sections


def row_to_dict(header, row):
    return {h: (row[i] if i < len(row) else '') for i, h in enumerate(header)}


def norm_datadefine(dd):
    """raw 文档 DataDefine 为 {<br>...} HTML 换行形式 -> JSON 字符串;spec 已是 JSON 则透传"""
    if not dd:
        return None
    if isinstance(dd, (dict, list)):
        return json.dumps(dd, ensure_ascii=False)
    s = str(dd).replace('<br>', '\n').strip()
    try:
        json.loads(s)
        return s
    except Exception:
        return None


def ensure_refs_present(spec, idx):
    """引用完整性:引用服务 *Input、事件 *Output/*Condition 引用的点位,缺失时自动补入
    FromDeviceType 引用清单(引用优先,字段由类型兜底)。返回自动补充清单。"""
    added = []
    select = spec['select']
    add_ids = {r['*ID'] for dim in DIMS for r in spec['add'].get(dim, []) if r.get('*ID')}

    def in_model(iid):
        return (iid in select.get('Attribute', [])
                or iid in select.get('MeasurePoint', [])
                or iid in add_ids)

    def ensure(iid, src):
        if not iid or in_model(iid):
            return
        if iid in idx['MeasurePoint']:
            select.setdefault('MeasurePoint', []).append(iid)
            added.append((src, f'MeasurePoint/{iid}'))
        elif iid in idx['Attribute']:
            select.setdefault('Attribute', []).append(iid)
            added.append((src, f'Attribute/{iid}'))
        else:
            print(f'WARN: {src} 引用点位 {iid} 在类型四维度中不存在')

    for sid in select.get('Service', []):
        row = idx['Service'].get(sid, {})
        for ref in (row.get('*Input') or '').split(','):
            ensure(ref.strip(), f'服务 {sid} 的 *Input')
    for eid in select.get('Event', []):
        row = idx['Event'].get(eid, {})
        for ref in (row.get('*Output') or '').split(','):
            ensure(ref.strip(), f'事件 {eid} 的 *Output')
        cond = (row.get('*Condition') or '').strip()
        m = re.match(r'^\s*([A-Za-z0-9_]+)\s*=', cond)
        if m and m.group(1) != (row.get('*Output') or '').strip():
            ensure(m.group(1), f'事件 {eid} 的 *Condition')
    if added:
        print('自动补充引用(服务 Input / 事件 Output 依赖点位):')
        for src, tgt in added:
            print(f'  + {src} -> {tgt}')
    return added


def fill_sheet(ws, rows):
    """按表头列名映射填充(不硬编码列号);DataDefine 规范化"""
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    hdr = [c.value for c in ws[1]]
    for r_i, row in enumerate(rows, 2):
        for col, h in enumerate(hdr, 1):
            v = row.get(h, '')
            if h == 'DataDefine':
                v = norm_datadefine(v)
            elif v == '':
                v = None
            ws.cell(row=r_i, column=col, value=v)


def main():
    if len(sys.argv) < 2:
        print('用法: python gen_device_model.py <model_spec.json> [输出路径]')
        sys.exit(1)
    spec_path = sys.argv[1]
    spec = json.loads(io.open(spec_path, encoding='utf-8').read())

    raw_doc = spec.get('raw_doc', '/opt/data/wiki/raw/papers/public_ElectricMeter_3P_V1_0_2.md')
    template = spec.get('template') or os.path.join(BASE, 'templates', 'model_template.xlsx')
    output = sys.argv[2] if len(sys.argv) > 2 else spec.get('output')
    if not output:
        print('ERROR: 未指定输出路径(spec.output 或命令行参数)')
        sys.exit(1)

    secs = parse_raw(raw_doc)
    idx = {}
    for dim in DIMS:
        rows = secs.get(dim, [])
        if not rows:
            idx[dim] = {}
            continue
        hdr = rows[0]
        idx[dim] = {r[0]: row_to_dict(hdr, r) for r in rows[1:] if r}

    ensure_refs_present(spec, idx)

    def pick(dim, ids):
        out = []
        for i in ids:
            if i not in idx[dim]:
                print(f'WARN: {dim} 缺 {i}')
                continue
            out.append(idx[dim][i])
        return out

    ref_rows = {dim: pick(dim, spec['select'].get(dim, [])) for dim in DIMS}
    add_rows = {dim: spec['add'].get(dim, []) for dim in DIMS}
    print(f'引用(FromDeviceType): ' + '  '.join(f'{d} {len(ref_rows[d])}' for d in DIMS))
    print(f'新增(四张子表): ' + '  '.join(f'{d} {len(add_rows[d])}' for d in DIMS))

    wb = openpyxl.load_workbook(template)

    ws = wb['BasicInfo']
    hdr = [c.value for c in ws[1]]
    m = spec['model']
    vals = {'*ID': m['id'], '*Name_default': m['name'], 'Name_zh_CN': m.get('name_zh', m['name']),
            'Name_en_US': m.get('name_en', ''), '*DeviceType': m['device_type']}
    for col, h in enumerate(hdr, 1):
        ws.cell(row=2, column=col, value=vals.get(h))

    ws = wb['FromDeviceType']
    hdr = [c.value for c in ws[1]]
    vals = {d: ','.join(spec['select'].get(d, [])) for d in DIMS}
    for col, h in enumerate(hdr, 1):
        ws.cell(row=2, column=col, value=vals.get(h))

    for dim in DIMS:
        fill_sheet(wb[dim], add_rows[dim])

    os.makedirs(os.path.dirname(output), exist_ok=True)
    wb.save(output)
    print('saved:', output)


if __name__ == '__main__':
    main()
