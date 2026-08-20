#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成设备类型 Excel(分支 B)—— 基于说明书四维度点位新建私有设备类型

输入:
  type_spec.json —— 设备类型配置(类型元信息 + 四维度点位全集)

输出:
  按模板 type_template.xlsx 结构生成的设备类型 xlsx

用法:
  python gen_device_type.py <type_spec.json> [输出路径覆盖]

铁律:
  - 内容全部来自用户文档(spec.points),标识符规范化(优先文档已有英文 ID,
    否则按语义译 CamelCase)。
  - 生成物 ID 一律 project_ 前缀;私有类型不回灌公有知识库。
  - 模板 type_template.xlsx 5 sheet,无 FromDeviceType(类型不引用其他类型)。
  - 列序与模型模板不同(DataDefine 在前、R/W 在后),按表头列名映射填充。

type_spec.json 结构:
{
  "template": "templates/type_template.xlsx",
  "output": "/opt/data/output/xxx_设备类型_yyyyMMdd.xlsx",
  "type": {
    "id": "project_xxx", "name": "...", "name_zh": "...", "name_en": "...",
    "category": "NORMAL", "domain": "配电", "desc": "...", "parent_type": ""
  },
  "points": {
    "Attribute": [{"*ID": "...", "*Name_default": "...", "*DataType": "...",
                   "DataDefine": "...", "Unit": "...", "*IsRequired": "False", "Desc": "..."}],
    "MeasurePoint": [{"*ID": "...", ..., "*R/W": "R", "Unit": "...", "Desc": "..."}],
    "Event": [{"*ID": "...", ..., "*EventType": "...", "*Output": "...", "*Condition": "...", "Desc": "..."}],
    "Service": [{"*ID": "...", ..., "*Input": "...", "Output": "...", "Desc": "..."}]
  }
}
"""
import io, json, os, sys
import openpyxl

DIMS = ['Attribute', 'MeasurePoint', 'Event', 'Service']
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def norm_datadefine(dd):
    if not dd:
        return None
    if isinstance(dd, (dict, list)):
        return json.dumps(dd, ensure_ascii=False)
    s = str(dd).strip()
    try:
        json.loads(s)
        return s
    except Exception:
        return None


def fill_sheet(ws, rows):
    """按表头列名映射填充(不硬编码列号);清空模板自带示例行;DataDefine 规范化"""
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    hdr = [c.value for c in ws[1]]
    for r_i, row in enumerate(rows, 2):
        for col, h in enumerate(hdr, 1):
            v = row.get(h, '')
            if h == 'DataDefine':
                v = norm_datadefine(v)
            elif v == '':
                v = None
            ws.cell(row=r_i, column=col, value=v)


def main():
    if len(sys.argv) < 2:
        print('用法: python gen_device_type.py <type_spec.json> [输出路径]')
        sys.exit(1)
    spec_path = sys.argv[1]
    spec = json.loads(io.open(spec_path, encoding='utf-8').read())

    template = spec.get('template') or os.path.join(BASE, 'templates', 'type_template.xlsx')
    output = sys.argv[2] if len(sys.argv) > 2 else spec.get('output')
    if not output:
        print('ERROR: 未指定输出路径(spec.output 或命令行参数)')
        sys.exit(1)

    t = spec['type']
    if not t['id'].startswith('project_'):
        print(f"ERROR: 类型 ID 必须以 project_ 开头,当前: {t['id']!r}")
        sys.exit(1)

    wb = openpyxl.load_workbook(template)

    ws = wb['BasicInfo']
    hdr = [c.value for c in ws[1]]
    vals = {'*ID': t['id'], '*Name_default': t['name'], 'Name_zh_CN': t.get('name_zh', t['name']),
            'Name_en_US': t.get('name_en', ''), '*Category': t.get('category', 'NORMAL'),
            '*Domain': t.get('domain', ''), 'Desc': t.get('desc', ''),
            'ParentType': t.get('parent_type', '')}
    for col, h in enumerate(hdr, 1):
        ws.cell(row=2, column=col, value=vals.get(h))

    points = spec.get('points', {})
    for dim in DIMS:
        fill_sheet(wb[dim], points.get(dim, []))
        print(f'{dim}: {len(points.get(dim, []))} 个')

    os.makedirs(os.path.dirname(output), exist_ok=True)
    wb.save(output)
    print('saved:', output)


if __name__ == '__main__':
    main()
