#!/usr/bin/env python3
"""Schema 2.1 的生成与校验入口（流水线核心模块）。

本模块是整个技能流水线 v2 的实现层，提供：
  - 设备模型生成 generate_model（分支 A）
  - 设备类型生成 generate_type（分支 B）
  - 协议点表生成 generate_point_table（分支 A2）
  - 三类产物的校验 verify_model / verify_type / verify_point
  - 命令行入口 model_main / type_main / point_main / verify_main

gen_device_model.py / gen_device_type.py / gen_point_table.py / verify_output.py
四个脚本的顶层 if __name__=='__main__' 均委托给本模块的对应入口函数，
因此本模块是实际执行路径。各脚本内的 main() 为旧版独立实现，仅作参考。

所有生成函数在写入前先经 artifact_contract.validate_artifact 契约校验，
并通过 atomic_save_workbook / atomic_write_json 原子写入，失败不覆盖正式文件。
"""
from __future__ import annotations

import argparse
import io
import json
import os
import re
import sys
from pathlib import Path

import openpyxl

from artifact_contract import (
    ContractError,
    DIMS,
    PROTOCOL_PROFILES,
    SCHEMA_VERSION,
    atomic_save_workbook,
    atomic_write_json,
    catalog_path_for,
    normalize_datadefine,
    normalize_enum_datadefine,
    parse_datadefine,
    protocol_family,
    validate_artifact,
    validate_point_groups,
)
from workspace import resolve_out

# 工程根目录（scripts 的上一级）
BASE = Path(__file__).resolve().parent.parent
# 协议模板目录
TPL_DIR = BASE / "templates"
# 模型 *DataType（大写枚举）-> 点表 dataType（协议寄存器类型）的映射
DTYPE_MAP = {
    "FLOAT": "float32", "INT": "i32", "STRING": "str", "ENUM": "u16",
    "BOOL": "bool", "DATETIME": "str",
}
# Modbus 各 dataType 所需的最小寄存器宽度（用于校验 registerCount 是否足够）
REGISTER_WIDTH = {
    "u8": 1, "i8": 1, "u16": 1, "i16": 1, "bool": 1, "bits": 1,
    "u32": 2, "i32": 2, "float32": 2, "ieee754_f32": 2,
    "u64": 4, "i64": 4, "double64": 4, "ieee754_f64": 4,
}


def load_json(path):
    """读取 UTF-8 编码的 JSON 文件并返回解析后的 Python 对象。

    输入：path —— JSON 文件路径
    输出：解析后的 dict/list
    调用情况：被 model_main / type_main / point_main 调用，加载 spec 文件。
    """
    with io.open(path, encoding="utf-8") as handle:
        return json.load(handle)


def load_header(sheet):
    """读取 worksheet 第 1 行作为表头列表。

    输入：sheet —— openpyxl Worksheet 对象
    输出：[单元格值, ...]
    调用情况：被 rows_after_header / fill_sheet / 各 generate_* / verify_* 调用。
    """
    return [cell.value for cell in sheet[1]]


def rows_after_header(sheet):
    """读取表头之后的所有数据行，按列名映射为字典列表。

    输入：sheet —— openpyxl Worksheet 对象（第 1 行表头）
    输出：(header, rows)
        header —— 表头列表
        rows   —— [{列名: 值}, ...]，跳过首列为 None 的空行
    调用情况：被 verify_model / verify_type / verify_point 调用，读取已生成的 Excel 做校验。
    """
    header = load_header(sheet)
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not values or values[0] is None:
            continue
        rows.append({key: values[index] if index < len(values) else None
                     for index, key in enumerate(header) if key is not None})
    return header, rows


def parse_raw(path):
    """解析 raw 物模型 Markdown 文档为按维度索引的点位字典。

    输入：
      path —— wiki/raw/papers/ 下的公有类型源文档路径
    输出：
      {维度: {*ID: {列名: 值}}}
      如 {"Attribute": {"SN": {"*ID": "SN", "*Name_default": "序列号", ...}}}
    作用：
      - 按 "## Attribute/MeasurePoint/Event/Service" 二级标题切分章节
      - 收集表格行（跳过分隔线），首行作表头，其余行按 *ID 索引
      - 与 gen_device_model.parse_raw 类似，但输出结构为按 ID 索引的字典
    调用情况：被 generate_model 调用，构建公有类型点位索引。
    """
    with io.open(path, encoding="utf-8") as handle:
        text = handle.read()
    sections, current = {}, None
    for line in text.splitlines():
        match = re.match(r"^## (Attribute|MeasurePoint|Event|Service)$", line.strip())
        if match:
            current = match.group(1)
            sections[current] = []
        elif current and line.startswith("|") and not re.match(r"^\|\s*-", line):
            cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
            if cells and cells[0]:
                sections[current].append(cells)
    result = {}
    for dim in DIMS:
        table = sections.get(dim, [])
        if not table:
            result[dim] = {}
            continue
        header = table[0]
        result[dim] = {
            row[0]: {key: row[index] if index < len(row) else ""
                     for index, key in enumerate(header)}
            for row in table[1:] if row and row[0]
        }
    return result


def fill_sheet(sheet, rows):
    """按表头列名映射填充 sheet 数据行（不硬编码列号）。

    输入：
      sheet —— openpyxl Worksheet 对象（第 1 行为表头）
      rows  —— 行字典列表 [{列名: 值}, ...]
    输出：无（就地修改 worksheet）
    作用：
      - 先删除模板自带的示例数据行（保留表头）
      - 按表头列名从行字典取值填充
      - DataDefine 列经 normalize_datadefine 规范化（非法 JSON 会抛 ContractError）
      - 空字符串转为 None
    调用情况：被 generate_model / generate_type 调用，填充四张维度子表。
    """
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    header = load_header(sheet)
    for row_index, row in enumerate(rows, 2):
        for column, key in enumerate(header, 1):
            if key is None:
                continue
            value = row.get(key)
            if key == "DataDefine":
                value = normalize_datadefine(value, f"{sheet.title}/{row.get('*ID')}.DataDefine")
                # ENUM 点位写入时规整为平台导入要求的 mappingItemList + enumKeyCode 格式
                if row.get("*DataType") == "ENUM":
                    parsed = parse_datadefine(value, f"{sheet.title}/{row.get('*ID')}.DataDefine")
                    parsed = normalize_enum_datadefine(parsed)
                    value = None if parsed is None else json.dumps(parsed, ensure_ascii=False)
            sheet.cell(row=row_index, column=column, value=None if value == "" else value)


def _split_refs(value):
    """将逗号分隔的引用字符串拆分为 ID 列表。

    输入：value —— 如 "Ua,Ub,Uc" 或 None
    输出：["Ua", "Ub", "Uc"]（去空白、去空项）
    作用：事件 *Output、服务 *Input/Output 等字段用逗号分隔多个点位引用，
          此函数统一拆分逻辑。
    调用情况：被 ensure_model_references / verify_model 调用。
    """
    return [item.strip() for item in str(value or "").split(",") if item.strip()]


def _condition_ref(value):
    """从事件 *Condition 等式中提取左侧引用的点位 ID。

    输入：value —— 如 "AlarmFlag = 1" 或 None
    输出：左侧点位 ID（"AlarmFlag"），非等式格式返回 None
    作用：事件 *Condition 形如 "点位 = 值"，需提取左侧点位做引用校验。
    调用情况：被 ensure_model_references / verify_model 调用。
    """
    match = re.fullmatch(r"\s*([A-Za-z][A-Za-z0-9_]*)\s*=\s*(.+?)\s*", str(value or ""))
    return match.group(1) if match else None


def ensure_model_references(spec, public_index):
    """引用完整性补全：事件/服务引用的点位缺失时自动补入 select 引用清单。

    输入：
      spec          —— model_spec 字典（含 select 引用清单和 add 新增点位）
      public_index  —— 公有类型四维度点位索引（来自 parse_raw）
    输出：无（就地修改 spec.select，补入缺失引用）
    作用：
      - 事件的 *Output 和 *Condition 引用的点位必须存在于模型中
      - 服务的 *Input 和 Output 引用的点位必须存在于模型中
      - 缺失时优先从公有类型补入 select（引用优先），公有类型也没有则抛 ContractError
      - 与 gen_device_model.ensure_refs_present 类似，但缺失时直接报错而非 WARN
    调用情况：被 generate_model 调用。
    """
    selected = spec["select"]
    additions = spec["add"]
    # 模型中已有的点位 ID 全集（select 引用 + add 新增）
    available = {
        "Attribute": set(selected.get("Attribute", [])) |
                     {row.get("*ID") for row in additions.get("Attribute", [])},
        "MeasurePoint": set(selected.get("MeasurePoint", [])) |
                        {row.get("*ID") for row in additions.get("MeasurePoint", [])},
    }

    def ensure(ref, source):
        """确保引用点位存在于模型，缺失则从公有类型补入 select。"""
        if not ref or ref in available["Attribute"] or ref in available["MeasurePoint"]:
            return
        for dim in ("MeasurePoint", "Attribute"):
            if ref in public_index[dim]:
                selected.setdefault(dim, []).append(ref)
                available[dim].add(ref)
                print(f"自动补充引用: {source} -> {dim}/{ref}")
                return
        raise ContractError(f"{source} 引用 {ref} 不在 Attribute/MeasurePoint 中")

    # 遍历引用的事件和新增事件，补全 *Output / *Condition 依赖
    for source_dim in ("Event", "Service"):
        public_rows = [public_index[source_dim][pid] for pid in selected.get(source_dim, [])
                       if pid in public_index[source_dim]]
        for row in public_rows + additions.get(source_dim, []):
            if source_dim == "Event":
                for ref in _split_refs(row.get("*Output")):
                    ensure(ref, f"Event {row.get('*ID')} *Output")
                ensure(_condition_ref(row.get("*Condition")), f"Event {row.get('*ID')} *Condition")
            else:
                for field in ("*Input", "Output"):
                    for ref in _split_refs(row.get(field)):
                        ensure(ref, f"Service {row.get('*ID')} {field}")


def build_catalog(model, public_rows, private_rows):
    """构建模型 catalog（点位清单 JSON），记录每个点位的来源和标准字段。

    输入：
      model         —— 模型元信息 {"id": ..., "device_type": ...}
      public_rows   —— 引用点位行字典 {维度: [行, ...]}（来自公有类型）
      private_rows  —— 新增点位行字典 {维度: [行, ...]}（来自 spec.add）
    输出：
      model_catalog 字典，含 schema_version、artifact_type、model_id、device_type、
      dimensions（按维度列出所有点位，每点标注 source=public/private 及标准字段）
    作用：
      - catalog 是模型 Excel 的伴随产物，供点表生成器获取点位的标准名称、单位、数据类型
      - 点表生成时从 catalog 读取 pointName/unit/dataType，确保与模型一致
    调用情况：被 generate_model 调用，结果经 atomic_write_json 写入 .catalog.json。
    """
    dimensions = {}
    for dim in DIMS:
        items = []
        for source, rows in (("public", public_rows[dim]), ("private", private_rows[dim])):
            for row in rows:
                items.append({
                    "dimension": dim,
                    "source": source,
                    "id": row.get("*ID"),
                    "name": row.get("Name_zh_CN") or row.get("*Name_default") or row.get("*ID"),
                    "name_default": row.get("*Name_default") or "",
                    "name_en": row.get("Name_en_US") or "",
                    "data_type": row.get("*DataType") or "",
                    "unit": row.get("Unit") or "",
                    "fields": row,
                })
        dimensions[dim] = items
    return {
        "schema_version": SCHEMA_VERSION,
        "artifact_type": "model_catalog",
        "model_id": model["id"],
        "device_type": model["device_type"],
        "dimensions": dimensions,
    }


def generate_model(spec, output=None):
    """生成设备模型 Excel 及其 catalog（分支 A）。

    输入：
      spec   —— model_spec 字典（raw_doc / model / select / add）
      output —— 可选输出路径覆盖（缺省取 spec.output）
    输出：
      {"output": 模型xlsx路径, "catalog": catalog json路径, "catalog_data": catalog字典}
    作用：
      1. validate_artifact 契约校验（schema 版本、ID 前缀、select/add 互斥等）
      2. parse_raw 解析公有类型文档，ensure_model_references 补全引用依赖
      3. 校验 select 中的 ID 在公有类型存在、无重复、与 add 无重叠
      4. 加载模型模板，填充 BasicInfo / FromDeviceType / 四张子表
      5. build_catalog 构建 catalog，atomic_save_workbook + atomic_write_json 原子写入
    调用情况：被 model_main 调用。
    异常：契约校验或引用完整性失败时抛 ContractError。
    """
    for warning in validate_artifact(spec, "model_spec"):
        print("WARN:", warning)
    output = output or spec.get("output") or resolve_out(basename=f"{spec['model']['id']}.xlsx")
    public_index = parse_raw(spec["raw_doc"])
    ensure_model_references(spec, public_index)
    # 平台铁律:设备类型中已存在的点位 ID,不得自定义同名点位;
    # 若想在 add 中"重定义"某测点/属性,必须先确认它不在公有类型对应维度里
    for dim in ("Attribute", "MeasurePoint"):
        for row in spec["add"].get(dim, []):
            pid = row.get("*ID")
            if pid and pid in public_index[dim]:
                raise ContractError(
                    f"model_spec.add.{dim} 点位 {pid} 已在设备类型 {spec['model']['device_type']} "
                    f"中存在,不得自定义同名点位;如需该点位请从类型 select 引用(接受类型定义的方向/字段),"
                    f"方向不符时用私有 Service 补下发能力"
                )
    # 从公有类型索引中按 select 清单提取引用行，并校验存在性与唯一性
    public_rows = {}
    for dim in DIMS:
        ids = spec["select"].get(dim, [])
        duplicates = sorted({pid for pid in ids if ids.count(pid) > 1})
        if duplicates:
            raise ContractError(f"model_spec.select.{dim} 存在重复 ID: {duplicates}")
        missing = [pid for pid in ids if pid not in public_index[dim]]
        if missing:
            raise ContractError(f"model_spec.select.{dim} 在公有类型中不存在: {missing}")
        public_rows[dim] = [public_index[dim][pid] for pid in ids]
    private_rows = {dim: spec["add"].get(dim, []) for dim in DIMS}
    # 校验 select 与 add 互斥（同一 ID 不能既引用又新增）
    for dim in DIMS:
        overlap = set(spec["select"].get(dim, [])) & {
            row.get("*ID") for row in private_rows[dim] if row.get("*ID")
        }
        if overlap:
            raise ContractError(f"{dim} 同时出现在 select 和 add: {sorted(overlap)}")

    # 加载模型模板并校验必需 sheet 齐全
    workbook = openpyxl.load_workbook(spec.get("template") or TPL_DIR / "model_template.xlsx")
    required_sheets = ["BasicInfo", "FromDeviceType", *DIMS]
    missing_sheets = [name for name in required_sheets if name not in workbook.sheetnames]
    if missing_sheets:
        raise ContractError(f"模型模板缺少 sheet: {missing_sheets}")
    # 填充 BasicInfo：模型元信息
    model = spec["model"]
    basic = workbook["BasicInfo"]
    values = {
        "*ID": model["id"], "*Name_default": model["name"],
        "Name_zh_CN": model.get("name_zh", model["name"]),
        "Name_en_US": model.get("name_en", ""), "*DeviceType": model["device_type"],
    }
    for column, key in enumerate(load_header(basic), 1):
        basic.cell(row=2, column=column, value=values.get(key))
    # 填充 FromDeviceType：四维度引用 ID 清单（逗号分隔）
    source_sheet = workbook["FromDeviceType"]
    for column, key in enumerate(load_header(source_sheet), 1):
        source_sheet.cell(row=2, column=column,
                          value=",".join(spec["select"].get(key, [])) if key in DIMS else None)
    # 填充四张子表：仅写入新增点位
    for dim in DIMS:
        fill_sheet(workbook[dim], private_rows[dim])

    # 构建 catalog 并原子写入模型 Excel 和 catalog JSON
    catalog = build_catalog(model, public_rows, private_rows)
    catalog_path = spec.get("model_catalog") or catalog_path_for(output)
    atomic_save_workbook(workbook, output)
    atomic_write_json(catalog, catalog_path)
    return {"output": str(output), "catalog": catalog_path, "catalog_data": catalog}


def generate_type(spec, output=None):
    """生成私有设备类型 Excel（分支 B）。

    输入：
      spec   —— type_spec 字典（type / points）
      output —— 可选输出路径覆盖（缺省取 spec.output）
    输出：
      {"output": 类型xlsx路径}
    作用：
      1. validate_artifact 契约校验（schema 版本、ID 前缀、点位分组与引用完整性）
      2. 加载类型模板，校验必需 sheet 齐全且无 FromDeviceType
      3. 填充 BasicInfo（类型元信息）和四张维度子表（全部来自 spec.points）
      4. atomic_save_workbook 原子写入
    调用情况：被 type_main 调用。
    """
    for warning in validate_artifact(spec, "type_spec"):
        print("WARN:", warning)
    output = output or spec.get("output") or resolve_out(basename=f"{spec['type']['id']}.xlsx")
    workbook = openpyxl.load_workbook(spec.get("template") or TPL_DIR / "type_template.xlsx")
    required_sheets = ["BasicInfo", *DIMS]
    missing_sheets = [name for name in required_sheets if name not in workbook.sheetnames]
    if missing_sheets:
        raise ContractError(f"类型模板缺少 sheet: {missing_sheets}")
    # 填充 BasicInfo：类型元信息
    type_info = spec["type"]
    values = {
        "*ID": type_info["id"], "*Name_default": type_info["name"],
        "Name_zh_CN": type_info.get("name_zh", type_info["name"]),
        "Name_en_US": type_info.get("name_en", ""),
        "*Category": type_info.get("category", "NORMAL"),
        "*Domain": type_info.get("domain", ""), "Desc": type_info.get("desc", ""),
        "ParentType": type_info.get("parent_type", ""),
    }
    for column, key in enumerate(load_header(workbook["BasicInfo"]), 1):
        workbook["BasicInfo"].cell(row=2, column=column, value=values.get(key))
    # 填充四张维度子表：内容全部来自 spec.points
    for dim in DIMS:
        fill_sheet(workbook[dim], spec["points"].get(dim, []))
    atomic_save_workbook(workbook, output)
    return {"output": str(output)}


def load_catalog(path):
    """加载并校验模型 catalog JSON 文件。

    输入：path —— .catalog.json 文件路径
    输出：catalog 字典
    作用：
      - 校验 schema_version 和 artifact_type 是否为 model_catalog
      - 校验 dimensions 字段存在且为 dict
    调用情况：被 generate_point_table / verify_point 调用，获取模型点位标准字段。
    异常：版本或类型不符时抛 ContractError。
    """
    catalog = load_json(path)
    if catalog.get("schema_version") != SCHEMA_VERSION or catalog.get("artifact_type") != "model_catalog":
        raise ContractError(f"模型 catalog 必须是 schema {SCHEMA_VERSION} model_catalog: {path}")
    dimensions = catalog.get("dimensions")
    if not isinstance(dimensions, dict):
        raise ContractError("模型 catalog 缺少 dimensions")
    return catalog


def normalize_address(value):
    """寄存器地址统一为十进制 int（支持十六进制字符串自动转换）。

    输入：value —— int、十进制字符串、或 "0x..." 十六进制字符串
    输出：十进制 int；空值返回 None
    调用情况：被 validate_protocol_rows / generate_point_table 调用。
    异常：非数字字符串抛 ValueError。
    """
    if value in (None, ""):
        return None
    if isinstance(value, str):
        value = value.strip()
        return int(value, 16) if re.fullmatch(r"0[xX][0-9a-fA-F]+", value) else int(value)
    return int(value)


def normalize_mask(value):
    """规范化 Modbus mask 字段为整数位掩码。

    输入：
      value —— 可为正整数、十六进制字符串、或位索引数组（如 "[0,3]" 表示 bit0|bit3）
    输出：
      int 位掩码（如 [0,3] -> 0b1001 = 9）
    作用：
      - 位索引数组 [0,3] 转为 sum(1<<bit) = 9
      - 字符串先尝试 JSON 解析为数组，否则按 int(str, 0) 解析（支持 0x 前缀）
      - 结果必须为正整数
    调用情况：被 validate_protocol_rows 调用，校验 mask 合法性并用于地址重叠判断。
    异常：非法值抛 ValueError。
    """
    if value in (None, ""):
        return None
    parsed = json.loads(value) if isinstance(value, str) and value.lstrip().startswith("[") else value
    if isinstance(parsed, list):
        if not parsed or any(not isinstance(bit, int) or isinstance(bit, bool) or bit < 0 for bit in parsed):
            raise ValueError
        return sum(1 << bit for bit in set(parsed))
    result = int(str(parsed), 0) if isinstance(parsed, str) else int(parsed)
    if result <= 0:
        raise ValueError
    return result


def validate_protocol_rows(protocol, rows, model_map=None):
    """校验协议点表行的语义正确性（必填字段、Modbus 宽度/功能码/地址重叠等）。

    输入：
      protocol   —— 协议名（如 "ModbusTCP_Vega_ARM64_V1.1.0"）
      rows       —— 点表行字典 {pointKey: {address, registerCount, functionCode, ...}}
      model_map  —— 可选，模型测点映射 {pointKey: {rw, ...}}（用于校验功能码与读写权限匹配）
    输出：
      errors —— 错误信息列表（空列表表示通过）
    作用：
      - 检查每行必填字段（按协议族 profile 的 required）
      - Modbus 专项校验：
        * address 为有效十进制/十六进制整数
        * registerCount 为正整数且 >= dataType 所需最小宽度
        * functionCode 与模型读写权限（R/W/RW）匹配
        * mask 为正整数/十六进制/位索引数组
        * map 为合法 JSON 对象或数组
        * 寄存器地址范围不重叠（同地址不同 mask 且 mask 不冲突时允许）
    调用情况：被 generate_point_table / verify_point 调用。
    """
    family = protocol_family(protocol)
    errors = []
    required = PROTOCOL_PROFILES[family]["required"]
    ranges = []
    for pid, row in rows.items():
        # 必填字段检查
        for field in required:
            if row.get(field) in (None, ""):
                errors.append(f"{pid} 缺少 {family} 必填字段 {field}")
        # 非 Modbus 协议只做必填字段检查
        if family != "modbus":
            continue
        # Modbus address 校验
        try:
            address = normalize_address(row.get("address"))
        except (TypeError, ValueError):
            errors.append(f"{pid} address 不是有效十进制/十六进制整数: {row.get('address')!r}")
            continue
        # registerCount 校验：必须为正整数
        count = row.get("registerCount")
        if not isinstance(count, int) or isinstance(count, bool) or count <= 0:
            errors.append(f"{pid} registerCount 必须是正整数")
            continue
        # dataType 宽度校验：registerCount 须 >= dataType 所需最小寄存器数
        dtype = str(row.get("dataType") or "").lower()
        minimum = REGISTER_WIDTH.get(dtype)
        if minimum and count < minimum:
            errors.append(f"{pid} {dtype} 至少需要 {minimum} 个寄存器, 当前 {count}")
        # functionCode 校验：解析功能码，校验与模型读写权限匹配
        code_text = str(row.get("functionCode") or "").upper()
        codes = re.findall(r"\b(?:0?[1-6]|0?F|10)\b", code_text)
        mode = (model_map or {}).get(pid, {}).get("rw", "R")
        has_read = any(code.lstrip("0") in {"1", "2", "3", "4"} for code in codes)
        has_write = any(code.lstrip("0") in {"5", "6", "F", "10"} for code in codes)
        if "R" in mode and not has_read:
            errors.append(f"{pid} 模型为 {mode}, functionCode 缺少读功能码")
        if "W" in mode and not has_write:
            errors.append(f"{pid} 模型为 {mode}, functionCode 缺少写功能码")
        if mode == "R" and has_write:
            errors.append(f"{pid} 模型为只读, functionCode 不应包含写功能码")
        # mask 校验：正整数、十六进制字符串或位索引数组
        mask = row.get("mask")
        normalized_mask = None
        if mask not in (None, ""):
            try:
                normalized_mask = normalize_mask(mask)
            except (TypeError, ValueError, json.JSONDecodeError):
                errors.append(f"{pid} mask 必须是正整数、十六进制字符串或非负位索引数组")
        # map 校验：合法 JSON 对象或数组
        mapping = row.get("map")
        if mapping not in (None, ""):
            try:
                parsed = json.loads(mapping) if isinstance(mapping, str) else mapping
                if not isinstance(parsed, (dict, list)):
                    raise ValueError
            except (json.JSONDecodeError, ValueError, TypeError):
                errors.append(f"{pid} map 必须是 JSON 对象或数组")
        if address is not None:
            ranges.append((address, address + count - 1, pid, normalized_mask))
    # 寄存器地址范围重叠检查（同地址不同 mask 且 mask 不冲突时允许共享）
    for index, first in enumerate(ranges):
        for second in ranges[index + 1:]:
            if first[0] <= second[1] and second[0] <= first[1]:
                shared_mask = (
                    first[0] == first[1] == second[0] == second[1]
                    and first[3] is not None and second[3] is not None
                    and first[3] & second[3] == 0
                )
                if not shared_mask:
                    errors.append(f"寄存器范围重叠: {first[2]}[{first[0]}-{first[1]}] 与 "
                                  f"{second[2]}[{second[0]}-{second[1]}]")
    return errors


def _catalog_measure_map(catalog):
    """从模型 catalog 提取 MeasurePoint 的标准字段映射。

    输入：catalog —— model_catalog 字典（来自 load_catalog）
    输出：
      {pointKey: {"pointName": ..., "unit": ..., "dtype": ..., "rw": ...}}
    作用：
      供点表生成器获取每个测点的标准名称、单位、数据类型和读写权限，
      确保点表与模型 catalog 一致。
    调用情况：被 generate_point_table / verify_point 调用。
    """
    result = {}
    for item in catalog["dimensions"].get("MeasurePoint", []):
        pid = item.get("id")
        if not pid:
            continue
        fields = item.get("fields") or {}
        result[pid] = {
            "pointName": item.get("name") or pid,
            "unit": item.get("unit") or "",
            "dtype": item.get("data_type") or "",
            "rw": fields.get("*R/W") or "R",
        }
    return result


def generate_point_table(spec, output=None):
    """生成协议点表 Excel（分支 A2）。

    输入：
      spec   —— point_reg 字典（protocol / model_xlsx / rows）
      output —— 可选输出路径覆盖（缺省取 spec.output）
    输出：
      {"output": 路径, "rows": 行数, "todo": 未生成测点列表, "inferred": dataType推断列表}
    作用：
      1. validate_artifact 契约校验（协议必填字段、pointKey 命名）
      2. 加载模型 catalog，获取测点标准字段
      3. 校验 spec.rows 的 pointKey 都在模型中，unit 与 catalog 一致
      4. dataType 缺省时按模型推断
      5. validate_protocol_rows 协议语义校验（Modbus 宽度/功能码/地址重叠等）
      6. 按协议模板表头列名填充，atomic_save_workbook 原子写入
      7. 无协议定位字段的测点进 todo 清单，不生成行
    调用情况：被 point_main 调用。
    异常：契约校验或协议语义校验失败时抛 ContractError。
    """
    for warning in validate_artifact(spec, "point_reg"):
        print("WARN:", warning)
    protocol = spec["protocol"]
    output = (output or spec.get("output")
              or resolve_out(basename=f"{Path(spec['model_xlsx']).stem}_{protocol}.xlsx"))
    template = spec.get("template") or TPL_DIR / f"{protocol}.xlsx"
    if not Path(template).is_file():
        raise ContractError(f"协议模板不存在: {template}")
    # 加载模型 catalog，获取测点标准字段
    catalog_path = spec.get("model_catalog") or catalog_path_for(spec["model_xlsx"])
    if not Path(catalog_path).is_file():
        raise ContractError(f"模型 catalog 不存在: {catalog_path}; 请先用 gen_device_model.py 重新生成模型")
    catalog = load_catalog(catalog_path)
    model_map = _catalog_measure_map(catalog)
    # 校验 spec.rows 的 pointKey 都在模型 catalog 中
    unknown = sorted(set(spec["rows"]) - set(model_map))
    if unknown:
        raise ContractError(f"point_reg.rows 含非模型测点: {unknown}")
    # 规范化每行：dataType 缺省推断、unit 一致性校验
    normalized_rows, inferred = {}, []
    for pid, reg in spec["rows"].items():
        normalized = dict(reg)
        if normalized.get("dataType") in (None, ""):
            normalized["dataType"] = DTYPE_MAP.get(model_map[pid]["dtype"])
            if normalized["dataType"]:
                inferred.append(pid)
        if "unit" in normalized and (normalized.get("unit") or "") != model_map[pid]["unit"]:
            raise ContractError(
                f"{pid} unit {normalized.get('unit')!r} 与模型 catalog {model_map[pid]['unit']!r} 不一致"
            )
        normalized_rows[pid] = normalized
    # 协议语义校验（必填字段、Modbus 宽度/功能码/地址重叠等）
    errors = validate_protocol_rows(protocol, normalized_rows, model_map)
    if errors:
        raise ContractError("\n".join(errors))

    # 按协议族定位字段判断哪些测点生成行
    family = protocol_family(protocol)
    locator = PROTOCOL_PROFILES[family]["locator"]
    output_rows, todo = [], []
    for pid, model_point in model_map.items():
        reg = normalized_rows.get(pid)
        if reg is None or reg.get(locator) in (None, ""):
            # 无协议定位信息（如 Modbus 无 address），进 todo 不生成行
            todo.append(pid)
            continue
        row = dict(reg)
        row.update({
            "pointName": model_point["pointName"], "pointKey": pid,
            "unit": model_point["unit"], "dataType": reg.get("dataType"),
        })
        if family == "modbus":
            row["address"] = normalize_address(reg["address"])
        output_rows.append(row)

    # 加载协议模板，按表头列名填充
    workbook = openpyxl.load_workbook(template)
    sheet = workbook.active
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    header = load_header(sheet)
    # 单列占位模板扩展为协议标准列集
    if header == ["pointName"]:
        header = list(PROTOCOL_PROFILES[family]["columns"])
        for column, key in enumerate(header, 1):
            sheet.cell(row=1, column=column, value=key)
    if not {"pointName", "pointKey"}.issubset(header):
        raise ContractError("点表模板必须包含 pointName 和 pointKey 列")
    missing_columns = sorted(set(PROTOCOL_PROFILES[family]["required"]) - set(header))
    if missing_columns:
        raise ContractError(f"点表模板缺少 {family} 必填列: {missing_columns}")
    # 列名别名修正（如模板中 efficient 实际对应 coefficient）
    aliases = {"efficient": "coefficient"}
    for row_index, row in enumerate(output_rows, 2):
        for column, key in enumerate(header, 1):
            if key is not None:
                value = row.get(aliases.get(key, key))
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
                sheet.cell(row=row_index, column=column, value=value)
    atomic_save_workbook(workbook, output)
    return {"output": str(output), "rows": len(output_rows), "todo": todo, "inferred": inferred}


def verify_model(path, errors, warnings, catalog_path=None):
    """校验已生成的设备模型 Excel（分支 A 产物）。

    输入：
      path         —— 设备模型 Excel 路径
      errors       —— 错误收集列表（就地追加）
      warnings     —— 警告收集列表（就地追加）
      catalog_path —— 可选，模型 catalog JSON 路径（缺省按模型路径推导）
    输出：无（通过 errors/warnings 副作用返回）
    作用：
      - 校验必需 sheet 齐全（BasicInfo/FromDeviceType/四维度）
      - 校验 BasicInfo 的 *ID 以 project_ 开头、*DeviceType 以 public_ 开头
      - 校验 FromDeviceType 引用与四张子表新增无重叠（互斥）
      - 校验新增点位分组结构（validate_point_groups）
      - 校验事件/服务引用的点位存在于模型全集
      - 校验 catalog 与模型四维度 ID 一致
    调用情况：被 verify_main 在 kind=model 时调用。
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    required = ["BasicInfo", "FromDeviceType", *DIMS]
    missing = [name for name in required if name not in workbook.sheetnames]
    if missing:
        errors.append(f"缺少 sheet: {missing}")
        return
    # 校验 BasicInfo 的 ID 前缀
    header = load_header(workbook["BasicInfo"])
    values = {key: workbook["BasicInfo"].cell(2, index + 1).value for index, key in enumerate(header)}
    if not str(values.get("*ID") or "").startswith("project_"):
        errors.append("BasicInfo *ID 必须以 project_ 开头")
    if not str(values.get("*DeviceType") or "").startswith("public_"):
        errors.append("BasicInfo *DeviceType 必须以 public_ 开头")
    # 读取 FromDeviceType 引用清单
    source_header = load_header(workbook["FromDeviceType"])
    selected = {dim: set(_split_refs(workbook["FromDeviceType"].cell(
        2, source_header.index(dim) + 1).value)) if dim in source_header else set() for dim in DIMS}
    # 读取四张子表新增点位
    additions = {}
    for dim in DIMS:
        _, additions[dim] = rows_after_header(workbook[dim])
    # 校验新增点位分组结构（必填字段、ID 命名、枚举、事件等式等）
    try:
        validate_point_groups(additions, "workbook")
    except ContractError as exc:
        errors.extend(str(exc).splitlines())
    # 校验引用与新增互斥
    for dim in DIMS:
        add_ids = [row.get("*ID") for row in additions[dim] if row.get("*ID")]
        overlap = selected[dim] & set(add_ids)
        if overlap:
            errors.append(f"{dim} 引用和新增重复: {sorted(overlap)}")
    # 校验事件/服务引用的点位存在于模型全集
    point_ids = (
        selected["Attribute"] | selected["MeasurePoint"] |
        {row.get("*ID") for dim in ("Attribute", "MeasurePoint") for row in additions[dim]}
    )
    for row in additions["Event"]:
        for ref in _split_refs(row.get("*Output")) + [_condition_ref(row.get("*Condition"))]:
            if ref and ref not in point_ids:
                errors.append(f"Event {row.get('*ID')} 引用不存在的点位 {ref}")
    for row in additions["Service"]:
        for field in ("*Input", "Output"):
            for ref in _split_refs(row.get(field)):
                if ref not in point_ids:
                    errors.append(f"Service {row.get('*ID')} {field} 引用不存在的点位 {ref}")
    # 校验 catalog 与模型四维度 ID 一致
    catalog_path = catalog_path or catalog_path_for(path)
    if not Path(catalog_path).is_file():
        errors.append(f"缺少模型 catalog: {catalog_path}")
    else:
        try:
            catalog = load_catalog(catalog_path)
            for dim in DIMS:
                catalog_ids = {item.get("id") for item in catalog["dimensions"].get(dim, [])}
                expected = selected[dim] | {row.get("*ID") for row in additions[dim]}
                if catalog_ids != expected:
                    errors.append(f"catalog {dim} ID 与模型不一致")
        except (ContractError, OSError, json.JSONDecodeError) as exc:
            errors.append(str(exc))


def verify_type(path, errors, warnings):
    """校验已生成的设备类型 Excel（分支 B 产物）。

    输入：
      path     —— 设备类型 Excel 路径
      errors   —— 错误收集列表（就地追加）
      warnings —— 警告收集列表（就地追加）
    输出：无（通过 errors 副作用返回）
    作用：
      - 校验必需 sheet 齐全（BasicInfo/四维度）且无 FromDeviceType
      - 校验 BasicInfo 的 *ID 以 project_ 开头
      - 校验四维度点位分组结构（validate_point_groups）
    调用情况：被 verify_main 在 kind=type 时调用。
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    required = ["BasicInfo", *DIMS]
    missing = [name for name in required if name not in workbook.sheetnames]
    if missing:
        errors.append(f"缺少 sheet: {missing}")
        return
    # 设备类型不应有 FromDeviceType（类型不引用其他类型）
    if "FromDeviceType" in workbook.sheetnames:
        errors.append("设备类型不应包含 FromDeviceType sheet")
    # 校验 BasicInfo 的 ID 前缀
    header = load_header(workbook["BasicInfo"])
    values = {key: workbook["BasicInfo"].cell(2, index + 1).value for index, key in enumerate(header)}
    if not str(values.get("*ID") or "").startswith("project_"):
        errors.append("BasicInfo *ID 必须以 project_ 开头")
    # 校验四维度点位分组结构
    groups = {dim: rows_after_header(workbook[dim])[1] for dim in DIMS}
    try:
        validate_point_groups(groups, "workbook")
    except ContractError as exc:
        errors.extend(str(exc).splitlines())


def verify_point(path, model_path, errors, warnings, catalog_path=None, protocol=None):
    """校验已生成的协议点表 Excel（分支 A2 产物）。

    输入：
      path         —— 点表 Excel 路径
      model_path   —— 对应的设备模型 Excel 路径（用于推导 catalog 路径）
      errors       —— 错误收集列表（就地追加）
      warnings     —— 警告收集列表（就地追加）
      catalog_path —— 可选，模型 catalog JSON 路径（缺省按模型路径推导）
      protocol     —— 可选，协议名（提供时执行协议语义校验）
    输出：无（通过 errors 副作用返回）
    作用：
      - 校验点表含 pointName 和 pointKey 列
      - 从模型 catalog 获取测点标准字段
      - 校验 pointKey 无重复、都在模型 MeasurePoint 中
      - 校验 pointName 和 unit 与模型 catalog 一致
      - 提供 protocol 时执行 validate_protocol_rows 协议语义校验
    调用情况：被 verify_main 在 kind=point 时调用。
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    header, rows = rows_after_header(workbook.active)
    for key in ("pointName", "pointKey"):
        if key not in header:
            errors.append(f"点表缺少 {key} 列")
            return
    # 加载模型 catalog，获取测点标准字段
    catalog_path = catalog_path or catalog_path_for(model_path)
    try:
        model_map = _catalog_measure_map(load_catalog(catalog_path))
    except (ContractError, OSError, json.JSONDecodeError) as exc:
        errors.append(f"无法读取模型 catalog: {exc}")
        return
    # 逐行校验 pointKey 唯一性、存在性、名称/单位一致性
    seen = set()
    row_map = {}
    for index, row in enumerate(rows, 2):
        pid = row.get("pointKey")
        if not pid:
            errors.append(f"第 {index} 行 pointKey 为空")
            continue
        if pid in seen:
            errors.append(f"点表 pointKey 重复: {pid}")
        seen.add(pid)
        if pid not in model_map:
            errors.append(f"点表 pointKey 不在模型 MeasurePoint 中: {pid}")
            continue
        expected = model_map[pid]
        if row.get("pointName") != expected["pointName"]:
            errors.append(f"{pid} pointName 与模型 catalog 不一致")
        if (row.get("unit") or "") != expected["unit"]:
            errors.append(f"{pid} unit 与模型 catalog 不一致")
        row_map[pid] = row
    # 提供协议名时执行协议语义校验
    if protocol:
        errors.extend(validate_protocol_rows(protocol, row_map, model_map))


def model_main():
    """设备模型生成命令行入口（分支 A）。

    用法：python gen_device_model.py <model_spec.json> [输出路径]
    流程：加载 spec -> generate_model -> 打印输出路径和 catalog 路径。
    调用情况：被 gen_device_model.py 顶层 if __name__=='__main__' 委托调用。
    """
    if len(sys.argv) < 2:
        raise ContractError("用法: python gen_device_model.py <model_spec.json> [输出路径]")
    result = generate_model(load_json(sys.argv[1]), sys.argv[2] if len(sys.argv) > 2 else None)
    print("saved:", result["output"])
    print("catalog:", result["catalog"])


def type_main():
    """设备类型生成命令行入口（分支 B）。

    用法：python gen_device_type.py <type_spec.json> [输出路径]
    流程：加载 spec -> generate_type -> 打印输出路径。
    调用情况：被 gen_device_type.py 顶层 if __name__=='__main__' 委托调用。
    """
    if len(sys.argv) < 2:
        raise ContractError("用法: python gen_device_type.py <type_spec.json> [输出路径]")
    result = generate_type(load_json(sys.argv[1]), sys.argv[2] if len(sys.argv) > 2 else None)
    print("saved:", result["output"])


def point_main():
    """协议点表生成命令行入口（分支 A2）。

    用法：python gen_point_table.py <point_reg.json> [输出路径]
    流程：加载 spec -> generate_point_table -> 打印输出路径、行数、todo、推断项。
    调用情况：被 gen_point_table.py 顶层 if __name__=='__main__' 委托调用。
    """
    if len(sys.argv) < 2:
        raise ContractError("用法: python gen_point_table.py <point_reg.json> [输出路径]")
    result = generate_point_table(load_json(sys.argv[1]), sys.argv[2] if len(sys.argv) > 2 else None)
    print(f"saved: {result['output']} ({result['rows']} rows)")
    if result["todo"]:
        print("未提供协议定位字段、不生成行:", ", ".join(result["todo"]))
    if result["inferred"]:
        print("dataType 推断:", ", ".join(result["inferred"]))


def verify_main():
    """交付物校验命令行入口（质量闸门）。

    用法：
      python verify_output.py --kind model  --xlsx <模型xlsx> [--catalog <catalog.json>]
      python verify_output.py --kind type   --xlsx <类型xlsx>
      python verify_output.py --kind point  --xlsx <点表xlsx> --model <模型xlsx>
                              [--catalog <catalog.json>] --protocol <协议名>
    输出：JSON 格式校验报告（passed/errors/warnings），退出码 0=通过 1=失败。
    调用情况：被 verify_output.py 顶层 if __name__=='__main__' 委托调用。
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("--kind", required=True, choices=("model", "type", "point"))
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--model")
    parser.add_argument("--catalog")
    parser.add_argument("--protocol")
    args = parser.parse_args()
    errors, warnings = [], []
    if not Path(args.xlsx).is_file():
        errors.append(f"文件不存在: {args.xlsx}")
    else:
        try:
            if args.kind == "model":
                verify_model(args.xlsx, errors, warnings, args.catalog)
            elif args.kind == "type":
                verify_type(args.xlsx, errors, warnings)
            elif not args.model:
                errors.append("kind=point 时必须提供 --model")
            elif not args.protocol:
                errors.append("kind=point 时必须提供 --protocol，以执行协议语义校验")
            else:
                verify_point(args.xlsx, args.model, errors, warnings, args.catalog, args.protocol)
        except (ContractError, KeyError, OSError, ValueError) as exc:
            errors.append(str(exc))
    report = {
        "kind": args.kind, "xlsx": args.xlsx, "passed": not errors,
        "errors": errors, "warnings": warnings,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if not errors else 1


def run(entry):
    """统一异常处理包装器：捕获契约/IO/JSON 异常并转为退出码。

    输入：
      entry —— 无参入口函数（如 model_main / type_main / point_main）
    输出：0=成功 1=失败
    作用：
      各生成脚本的顶层 if __name__=='__main__' 通过此函数调用入口，
      统一捕获 ContractError/KeyError/OSError/ValueError/json.JSONDecodeError，
      打印 ERROR 信息并返回退出码 1。
    调用情况：被 gen_device_model.py / gen_device_type.py / gen_point_table.py 调用。
    """
    try:
        entry()
    except (ContractError, KeyError, OSError, ValueError, json.JSONDecodeError) as exc:
        print(f"ERROR: {exc}")
        return 1
    return 0
