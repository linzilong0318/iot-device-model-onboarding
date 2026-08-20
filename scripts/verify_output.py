#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
交付物校验脚本(verify_output.py)—— 多 agent 流程的质量闸门
对生成的 Excel 做机器可检校验,失败即 fail-fast(exit 1),供校验 agent 调用。

支持三类产物:
  --kind model   设备模型 Excel(分支 A)
  --kind type    设备类型 Excel(分支 B)
  --kind point   点表 Excel(需 --model 指定对应设备模型 Excel 做一致性比对)

用法:
  python verify_output.py --kind model  --xlsx <模型xlsx>
  python verify_output.py --kind type   --xlsx <类型xlsx>
  python verify_output.py --kind point  --xlsx <点表xlsx> --model <模型xlsx>

退出码: 0=通过  1=有校验失败项
输出: JSON 格式的校验报告(同时打印到 stdout),含 passed/errors/warnings
"""
import argparse, json, os, re, sys
import openpyxl

DIMS = ['Attribute', 'MeasurePoint', 'Event', 'Service']


def load_hdr(ws):
    return [c.value for c in ws[1]]


def rows_after_header(ws):
    hdr = load_hdr(ws)
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        out.append({h: (r[i] if i < len(r) else None) for i, h in enumerate(hdr)})
    return hdr, out


def check(cond, errors, msg):
    if not cond:
        errors.append(msg)


def verify_model(path, errors, warnings):
    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    for s in ['BasicInfo', 'FromDeviceType', 'Attribute', 'MeasurePoint', 'Event', 'Service']:
        check(s in sheets, errors, f'缺少 sheet: {s}')

    ws = wb['BasicInfo']
    hdr = load_hdr(ws)
    vals = {h: ws.cell(row=2, column=i + 1).value for i, h in enumerate(hdr)}
    mid = vals.get('*ID') or ''
    check(mid.startswith('project_'), errors,
          f"BasicInfo *ID 必须以 project_ 开头,当前: {mid!r}")
    dt = vals.get('*DeviceType') or ''
    check(dt.startswith('public_'), errors,
          f"BasicInfo *DeviceType 必须以 public_ 开头(引用公有类型),当前: {dt!r}")

    ws = wb['FromDeviceType']
    hdr = load_hdr(ws)
    from_ids = set()
    for h in DIMS:
        v = ws.cell(row=2, column=hdr.index(h) + 1).value if h in hdr else None
        for s in (v or '').split(','):
            s = s.strip()
            if s:
                from_ids.add((h, s))

    add_ids = set()
    for dim in DIMS:
        ws = wb[dim]
        _, rows = rows_after_header(ws)
        for r in rows:
            pid = r.get('*ID')
            if pid:
                add_ids.add((dim, pid))

    overlap = from_ids & add_ids
    check(not overlap, errors,
          f"FromDeviceType 引用点位重复出现在子表(应互斥): {sorted(overlap)[:10]}")

    for dim in DIMS:
        ws = wb[dim]
        _, rows = rows_after_header(ws)
        for r in rows:
            dd = r.get('DataDefine')
            if dd and isinstance(dd, str) and dd.strip():
                try:
                    json.loads(dd)
                except Exception as e:
                    errors.append(f"{dim}/{r.get('*ID')} DataDefine 非合法 JSON: {e}")

    ws = wb['Service']
    _, sv_rows = rows_after_header(ws)
    ws = wb['Event']
    _, ev_rows = rows_after_header(ws)
    all_model_ids = {pid for _, pid in from_ids} | {pid for _, pid in add_ids}

    for r in sv_rows:
        for ref in (r.get('*Input') or '').split(','):
            ref = ref.strip()
            if ref and ref not in all_model_ids:
                errors.append(f"Service {r.get('*ID')} 的 *Input 引用 {ref} 不在模型点位全集")
    for r in ev_rows:
        for ref in (r.get('*Output') or '').split(','):
            ref = ref.strip()
            if ref and ref not in all_model_ids:
                errors.append(f"Event {r.get('*ID')} 的 *Output 引用 {ref} 不在模型点位全集")
        cond = (r.get('*Condition') or '').strip()
        m = re.match(r'^\s*([A-Za-z0-9_]+)\s*=', cond)
        if m and m.group(1) not in all_model_ids:
            errors.append(f"Event {r.get('*ID')} 的 *Condition 引用 {m.group(1)} 不在模型点位全集")


def verify_type(path, errors, warnings):
    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    for s in ['BasicInfo', 'Attribute', 'MeasurePoint', 'Event', 'Service']:
        check(s in sheets, errors, f'缺少 sheet: {s}')
    check('FromDeviceType' not in sheets, errors,
          '设备类型模板不应有 FromDeviceType sheet')

    ws = wb['BasicInfo']
    hdr = load_hdr(ws)
    vals = {h: ws.cell(row=2, column=i + 1).value for i, h in enumerate(hdr)}
    tid = vals.get('*ID') or ''
    check(tid.startswith('project_'), errors,
          f"BasicInfo *ID 必须以 project_ 开头,当前: {tid!r}")

    for dim in DIMS:
        ws = wb[dim]
        _, rows = rows_after_header(ws)
        for r in rows:
            pid = r.get('*ID')
            check(bool(pid), errors, f"{dim} 存在 *ID 为空的行")
            dd = r.get('DataDefine')
            if dd and isinstance(dd, str) and dd.strip():
                try:
                    json.loads(dd)
                except Exception as e:
                    errors.append(f"{dim}/{pid} DataDefine 非合法 JSON: {e}")


def verify_point(path, model_path, errors, warnings):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    hdr = load_hdr(ws)
    check('pointKey' in hdr, errors, '点表缺少 pointKey 列')
    check('pointName' in hdr, errors, '点表缺少 pointName 列')
    _, rows = rows_after_header(ws)

    for idx, r in enumerate(rows, 2):
        pk = r.get('pointKey')
        check(bool(pk), errors, f'第 {idx} 行 pointKey 为空')

    if 'address' in hdr:
        for idx, r in enumerate(rows, 2):
            addr = r.get('address')
            if addr is not None and addr != '':
                check(isinstance(addr, int), errors,
                      f"第 {idx} 行 address 必须为十进制 int,当前: {addr!r}({type(addr).__name__})")

    mwb = openpyxl.load_workbook(model_path)
    mws = mwb['MeasurePoint']
    mhdr = load_hdr(mws)
    _, mrows = rows_after_header(mws)
    model_keys = {r.get('*ID') for r in mrows if r.get('*ID')}
    from_ws = mwb['FromDeviceType']
    fhdr = load_hdr(from_ws)
    if 'MeasurePoint' in fhdr:
        v = from_ws.cell(row=2, column=fhdr.index('MeasurePoint') + 1).value
        for s in (v or '').split(','):
            s = s.strip()
            if s:
                model_keys.add(s)

    pt_keys = {r.get('pointKey') for r in rows if r.get('pointKey')}
    not_in_model = pt_keys - model_keys
    check(not not_in_model, errors,
          f"点表 pointKey 不在设备模型测点全集中: {sorted(not_in_model)[:10]}")

    model_name_map = {}
    for r in mrows:
        pid = r.get('*ID')
        if pid:
            model_name_map[pid] = r.get('Name_zh_CN') or r.get('*Name_default') or ''
    for idx, r in enumerate(rows, 2):
        pk = r.get('pointKey')
        pn = r.get('pointName')
        if pk in model_name_map:
            expected = model_name_map[pk]
            if expected and pn and pn != expected:
                warnings.append(f"第 {idx} 行 pointName({pn!r}) 与模型({expected!r}) 不一致")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--kind', required=True, choices=['model', 'type', 'point'])
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--model', help='设备模型 Excel(仅 kind=point 时需要)')
    args = ap.parse_args()

    errors, warnings = [], []
    if not os.path.exists(args.xlsx):
        errors.append(f'文件不存在: {args.xlsx}')
    else:
        if args.kind == 'model':
            verify_model(args.xlsx, errors, warnings)
        elif args.kind == 'type':
            verify_type(args.xlsx, errors, warnings)
        elif args.kind == 'point':
            if not args.model:
                errors.append('kind=point 时必须提供 --model')
            elif not os.path.exists(args.model):
                errors.append(f'模型文件不存在: {args.model}')
            else:
                verify_point(args.xlsx, args.model, errors, warnings)

    report = {
        'kind': args.kind,
        'xlsx': args.xlsx,
        'passed': len(errors) == 0,
        'errors': errors,
        'warnings': warnings,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    sys.exit(0 if report['passed'] else 1)


if __name__ == '__main__':
    main()
