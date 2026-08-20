#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成点表 Excel(分支 A2)—— 基于设备模型 Excel + 寄存器映射,按通信协议选模板

输入:
  point_reg.json —— 点表配置(协议 + 模型路径 + 寄存器映射 pointKey -> 寄存器信息)
  设备模型 Excel —— pointKey/pointName/unit 的来源(平台按名称+标识符映射)

输出:
  按协议模板(templates/<协议名>.xlsx)结构生成的点表 xlsx

用法:
  python gen_point_table.py <point_reg.json> [输出路径覆盖]

铁律:
  - pointName/pointKey/unit 与设备模型一致(从模型 Excel 读取),绝不从用户文档直抄。
  - address 必须十进制(支持 0x 十六进制自动转换);dataType 以 spec 为准,
    缺省时按模型数据类型推断并标注。
  - 以说明书为准:spec 中无寄存器信息的测点【不生成行】,进 todo 清单供总结告知用户。
  - 点表只收录测点(MeasurePoint)维度,属性/事件/服务不入表。
  - 按通信协议选模板:模板文件位于 templates/<协议名>.xlsx,各协议列定义不同,
    一律按目标模板表头列名映射填充,不硬编码列号。

point_reg.json 结构:
{
  "protocol": "ModbusTCP_Vega_ARM64_V1.1.0",
  "template": "templates/ModbusTCP_Vega_ARM64_V1.1.0.xlsx",
  "model_xlsx": "/opt/data/output/xxx_设备模型_yyyyMMdd.xlsx",
  "output": "/opt/data/output/xxx_点表_yyyyMMdd.xlsx",
  "rows": {
    "Ua": {"address": "0x0040", "registerCount": 2, "functionCode": "[03,00]",
           "dataType": "float32", "coefficient": null, "unit": "V"},
    "ClearE": {"address": "0x000D", "registerCount": 1, "functionCode": "[03,06]",
               "dataType": "u16"}
  }
}
说明:
  - rows 的 key 是模型 pointKey(与设备模型 MeasurePoint 的 *ID 一致)
  - address 支持 int 或 "0x..." 字符串(自动转十进制)
  - dataType 缺省时按模型 *DataType 推断(FLOAT->float32, INT->i32, ENUM->u16...)
  - rows 中不出现的 pointKey = 说明书无寄存器信息,进 todo 清单,不生成行
  - 各协议特有列(如 MQTT 的 topic/jsonpath、OPC UA 的 tag)直接在 rows[pointKey] 里
    按模板列名提供即可,脚本按表头列名透传
"""
import io, json, os, re, sys
import openpyxl

DIMS = ['Attribute', 'MeasurePoint', 'Event', 'Service']
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TPL_DIR = os.path.join(BASE, 'templates')

DTYPE_MAP = {
    'FLOAT': 'float32', 'INT': 'i32', 'STRING': 'str', 'ENUM': 'u16',
    'BOOL': 'bool', 'DATETIME': 'str',
}


def load_hdr(ws):
    return [c.value for c in ws[1]]


def load_model_measure_points(path):
    """读设备模型 Excel -> {pointKey: {pointName, unit, dtype}}
    pointKey 全集 = FromDeviceType 引用测点 + 子表 MeasurePoint 新增测点"""
    wb = openpyxl.load_workbook(path)
    out = {}

    ws = wb['FromDeviceType']
    hdr = load_hdr(ws)
    if 'MeasurePoint' in hdr:
        v = ws.cell(row=2, column=hdr.index('MeasurePoint') + 1).value
        from_ids = [s.strip() for s in (v or '').split(',') if s.strip()]
    else:
        from_ids = []

    ws = wb['MeasurePoint']
    hdr = load_hdr(ws)
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        row = {h: (r[i] if i < len(r) else None) for i, h in enumerate(hdr)}
        pid = row.get('*ID')
        if pid:
            out[pid] = {
                'pointName': row.get('Name_zh_CN') or row.get('*Name_default') or pid,
                'unit': row.get('Unit') or '',
                'dtype': row.get('*DataType') or '',
            }
    wb.close()

    for pid in from_ids:
        if pid not in out:
            out[pid] = {'pointName': pid, 'unit': '', 'dtype': ''}
    return out


def norm_address(v):
    """寄存器地址统一为十进制 int;十六进制 '0x…' 自动转换。返回 (值, 是否转换)"""
    if v is None or v == '':
        return None, False
    if isinstance(v, str):
        s = v.strip()
        if re.match(r'^0[xX][0-9a-fA-F]+$', s):
            return int(s, 16), True
        return int(s), False
    return int(v), False


def build_rows(model_mps, reg_rows):
    """合并模型测点 + 寄存器映射 -> (生成行, todo, inferred, hexed)
    生成行只含 spec.rows 中有寄存器信息的测点;无信息的进 todo"""
    out, todo, inferred, hexed = [], [], [], []
    for pid, mp in model_mps.items():
        reg = reg_rows.get(pid)
        if reg is None:
            todo.append(pid)
            continue
        addr, is_hex = norm_address(reg.get('address'))
        if is_hex:
            hexed.append(pid)
        if addr is None:
            todo.append(pid)
            continue
        dtype = reg.get('dataType')
        dtype_inferred = False
        if dtype is None or dtype == '':
            mapped = DTYPE_MAP.get(mp['dtype'])
            if mapped:
                dtype = mapped
                dtype_inferred = True
                inferred.append(pid)
        row = {
            'pointName': mp['pointName'],
            'pointKey': pid,
            'unit': reg.get('unit', mp['unit']),
            'dataType': dtype,
            'address': addr,
            'registerCount': reg.get('registerCount'),
            'functionCode': reg.get('functionCode'),
            'coefficient': reg.get('coefficient'),
            'parentKey': reg.get('parentKey'),
            'order': reg.get('order'),
            'mask': reg.get('mask'),
            'map': reg.get('map'),
            'basicValue': reg.get('basicValue'),
            'wait': reg.get('wait'),
        }
        for k, v in reg.items():
            if k not in row and v is not None:
                row[k] = v
        if dtype_inferred:
            row['_dtype_inferred'] = True
        out.append(row)
    return out, todo, inferred, hexed


def write_point_table(rows, template, output):
    """按协议模板表头列名映射填充;清空模板自带数据行"""
    wb = openpyxl.load_workbook(template)
    ws = wb.active
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    hdr = load_hdr(ws)
    for r_i, row in enumerate(rows, 2):
        for col, h in enumerate(hdr, 1):
            if h is None:
                continue
            v = row.get(h)
            ws.cell(row=r_i, column=col, value=v if v is not None else None)
    os.makedirs(os.path.dirname(output), exist_ok=True)
    wb.save(output)
    return hdr


def main():
    if len(sys.argv) < 2:
        print('用法: python gen_point_table.py <point_reg.json> [输出路径]')
        sys.exit(1)
    spec_path = sys.argv[1]
    spec = json.loads(io.open(spec_path, encoding='utf-8').read())

    protocol = spec.get('protocol', '')
    template = spec.get('template')
    if not template:
        if not protocol:
            print('ERROR: 未指定 protocol 或 template')
            sys.exit(1)
        template = os.path.join(TPL_DIR, f'{protocol}.xlsx')
    if not os.path.exists(template):
        print(f'ERROR: 协议模板不存在: {template}')
        print(f'可用协议: {sorted(f[:-5] for f in os.listdir(TPL_DIR) if f.endswith(".xlsx"))}')
        sys.exit(1)

    model_xlsx = spec['model_xlsx']
    if not os.path.exists(model_xlsx):
        print(f'ERROR: 设备模型文件不存在: {model_xlsx}')
        sys.exit(1)

    output = sys.argv[2] if len(sys.argv) > 2 else spec.get('output')
    if not output:
        print('ERROR: 未指定输出路径(spec.output 或命令行参数)')
        sys.exit(1)

    model_mps = load_model_measure_points(model_xlsx)
    reg_rows = spec.get('rows', {})
    out, todo, inferred, hexed = build_rows(model_mps, reg_rows)

    print(f'协议: {protocol or os.path.basename(template)}')
    print(f'模型测点: {len(model_mps)}')
    print(f'点表行: {len(out)}  (仅测点;属性/事件/服务不进入点表)')
    print(f'未找到地址、不生成行(需告知用户): {len(todo)}'
          + (': ' + ', '.join(todo[:20]) + (' …' if len(todo) > 20 else '') if todo else ''))
    print(f'寄存器地址十六进制→十进制转换: {len(hexed)}'
          + (': ' + ', '.join(hexed) if hexed else ''))
    print(f'dataType 推断(说明书未给): {len(inferred)}'
          + (': ' + ', '.join(inferred[:20]) + (' …' if len(inferred) > 20 else '') if inferred else ''))

    hdr = write_point_table(out, template, output)
    print(f'saved: {output} (表头列数: {len([h for h in hdr if h])})')


if __name__ == '__main__':
    main()
